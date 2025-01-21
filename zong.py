import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import glob


input_files = glob.glob("D:/Desktop/冬日赛季/data/*.json")  
output_file = "D:/Desktop/冬日赛季/EX/total_rankings_with_server_alliance.xlsx"  


total_data_df = pd.DataFrame()


for input_file in input_files:
    with open(input_file, "r", encoding="utf-8") as file:
        json_data = pd.read_json(file)

    if "Data" in json_data:
        data_df = pd.json_normalize(json_data["Data"])
    else:
        data_df = pd.DataFrame(json_data)

    
    data_df["kill_die_ratio"] = data_df["c_sumkill"] / data_df["c_die"]
    data_df["kill_die_ratio"] = data_df["kill_die_ratio"].replace([float("inf"), -float("inf")], 0).fillna(0)

    # 提取服务器和联盟名称（wid 和 gnick）
    data_df["wid"] = data_df["wid"].fillna("未知服务器")  
    data_df["gnick"] = data_df["gnick"].fillna("未知联盟")  

    # 合并到总数据框中
    total_data_df = pd.concat([total_data_df, data_df], ignore_index=True)

# 确保列名正确后再排序
if "c_die" in total_data_df.columns and "c_sumkill" in total_data_df.columns:
    # 按阵亡降序、击杀升序、击杀比降序排序
    sorted_df = total_data_df.sort_values(by=["c_die", "c_sumkill", "kill_die_ratio"], ascending=[False, True, False])

   
    columns_to_export = {
        "nick": "昵称",
        "wid": "服务器",
        "gnick": "联盟名称",
        "c_die": "阵亡",
        "c_sumkill": "击杀",
        "kill_die_ratio": "阵亡击杀比",
        "maxpower": "最高战力"
    }
    export_df = sorted_df[list(columns_to_export.keys())].rename(columns=columns_to_export)

   
    export_df.to_excel(output_file, index=False)

    
    workbook = load_workbook(output_file)
    sheet = workbook.active

    # 设置单元格样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            # 添加边框
            cell.border = thin_border
            # 居中对齐
            cell.alignment = Alignment(horizontal="center", vertical="center")

   
    workbook.save(output_file)

    print(f"总排行榜数据已成功导出到 {output_file}，并完成格式化。")
else:
    print("找不到列 'c_die' 或 'c_sumkill'，请检查 JSON 文件格式。")
