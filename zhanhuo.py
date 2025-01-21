import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


input_file = "D:/Desktop/8.json"  
output_file = "D:/Desktop/sorted_data.xlsx"  


with open(input_file, "r", encoding="utf-8") as file:
    json_data = pd.read_json(file)


if "Data" in json_data:
    data_df = pd.json_normalize(json_data["Data"])
else:
    data_df = pd.DataFrame(json_data)


data_df["kill_die_ratio"] = data_df["c_sumkill"] / data_df["c_die"]
data_df["kill_die_ratio"] = data_df["kill_die_ratio"].replace([float("inf"), -float("inf")], 0).fillna(0)


if "c_die" in data_df.columns and "c_sumkill" in data_df.columns:
    
    sorted_df = data_df.sort_values(by=["c_die", "c_sumkill", "kill_die_ratio"], ascending=[False, True, False])

    
    columns_to_export = {
        "nick": "昵称",
        "c_die": "阵亡",
        "c_sumkill": "击杀",
        "kill_die_ratio": "阵亡击杀比",
        "maxpower": "最高战力"
    }
    export_df = sorted_df[list(columns_to_export.keys())].rename(columns=columns_to_export)

    
    export_df.to_excel(output_file, index=False)

    
    workbook = load_workbook(output_file)
    sheet = workbook.active

    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            
            cell.border = thin_border
            
            cell.alignment = Alignment(horizontal="center", vertical="center")

   
    workbook.save(output_file)

    print(f"数据已成功导出到 {output_file}，并完成格式化。")
else:
    print("找不到列 'c_die' 或 'c_sumkill'，请检查 JSON 文件格式。")