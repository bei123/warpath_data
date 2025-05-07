import json
import requests
import argparse
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
import os
from pathlib import Path
import logging
import asyncio
import aiohttp
from concurrent.futures import ThreadPoolExecutor
from functools import partial
import time
from aiohttp import ClientTimeout, ContentTypeError
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class GuildDataFetcher:
    """公会数据获取器"""
    
    def __init__(self, base_url: str = "https://yx.dmzgame.com/warpath", max_concurrent: int = 10):
        self.base_url = base_url
        self.output_dir = Path("guild_data")
        self.pid_data_dir = Path("pid_data")  # 添加pid_data目录
        self.output_dir.mkdir(exist_ok=True)
        self.pid_data_dir.mkdir(exist_ok=True)  # 创建pid_data目录
        self.max_concurrent = max_concurrent
        self.session = None
        self.timeout = ClientTimeout(total=30)  # 30秒超时
        self.max_retries = 3  # 最大重试次数
        self.retry_delay = 2  # 重试延迟（秒）
    
    async def __aenter__(self):
        """异步上下文管理器入口"""
        self.session = aiohttp.ClientSession(timeout=self.timeout)
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """异步上下文管理器出口"""
        if self.session:
            await self.session.close()
    
    async def fetch_with_retry(self, url: str, params: Dict[str, Any], retry_count: int = 0) -> Optional[Dict[str, Any]]:
        """带重试机制的请求函数"""
        try:
            async with self.session.get(url, params=params) as response:
                if response.status == 200:
                    try:
                        return await response.json()
                    except ContentTypeError as e:
                        logging.warning(f"响应MIME类型不是JSON: {e}")
                        text = await response.text()
                        
                        # 尝试解析文本内容为JSON
                        try:
                            data = json.loads(text)
                            logging.info("成功解析响应内容为JSON")
                            return data
                        except json.JSONDecodeError:
                            logging.warning(f"响应内容不是有效的JSON: {text[:500]}...")
                            
                            # 检查是否是错误页面
                            if "error" in text.lower() or "exception" in text.lower():
                                logging.error(f"服务器返回错误页面: {text[:200]}...")
                                if retry_count < self.max_retries:
                                    wait_time = self.retry_delay * (retry_count + 1)
                                    logging.info(f"等待 {wait_time} 秒后重试...")
                                    await asyncio.sleep(wait_time)
                                    return await self.fetch_with_retry(url, params, retry_count + 1)
                                return None
                            
                            # 检查是否是维护页面
                            if "maintenance" in text.lower() or "维护" in text:
                                logging.error("服务器可能正在维护中")
                                if retry_count < self.max_retries:
                                    wait_time = self.retry_delay * (retry_count + 1)
                                    logging.info(f"等待 {wait_time} 秒后重试...")
                                    await asyncio.sleep(wait_time)
                                    return await self.fetch_with_retry(url, params, retry_count + 1)
                                return None
                            
                            # 其他HTML响应
                            if retry_count < self.max_retries:
                                wait_time = self.retry_delay * (retry_count + 1)
                                logging.info(f"等待 {wait_time} 秒后重试...")
                                await asyncio.sleep(wait_time)
                                return await self.fetch_with_retry(url, params, retry_count + 1)
                            return None
                elif response.status == 429:  # Too Many Requests
                    if retry_count < self.max_retries:
                        wait_time = self.retry_delay * (retry_count + 1)
                        logging.warning(f"请求过于频繁，等待 {wait_time} 秒后重试...")
                        await asyncio.sleep(wait_time)
                        return await self.fetch_with_retry(url, params, retry_count + 1)
                    return None
                else:
                    logging.error(f"请求失败，状态码: {response.status}")
                    text = await response.text()
                    logging.error(f"错误响应内容: {text[:500]}...")
                    if retry_count < self.max_retries:
                        wait_time = self.retry_delay * (retry_count + 1)
                        logging.info(f"等待 {wait_time} 秒后重试...")
                        await asyncio.sleep(wait_time)
                        return await self.fetch_with_retry(url, params, retry_count + 1)
                    return None
        except asyncio.TimeoutError:
            logging.error("请求超时")
            if retry_count < self.max_retries:
                wait_time = self.retry_delay * (retry_count + 1)
                logging.info(f"等待 {wait_time} 秒后重试...")
                await asyncio.sleep(wait_time)
                return await self.fetch_with_retry(url, params, retry_count + 1)
            return None
        except Exception as e:
            logging.error(f"请求发生错误: {e}")
            if retry_count < self.max_retries:
                wait_time = self.retry_delay * (retry_count + 1)
                logging.info(f"等待 {wait_time} 秒后重试...")
                await asyncio.sleep(wait_time)
                return await self.fetch_with_retry(url, params, retry_count + 1)
            return None
    
    async def fetch_guild_data(self, gid: int, day: str) -> Optional[Dict[str, Any]]:
        """异步获取公会成员数据"""
        url = f"{self.base_url}/guild_member"
        params = {"gid": gid, "day": day}
        
        data = await self.fetch_with_retry(url, params)
        if not data:
            logging.error(f"获取公会 {gid} 在 {day} 的数据失败")
        return data
    
    def extract_pids(self, data: Dict[str, Any]) -> List[int]:
        """提取pid列表"""
        if not isinstance(data, dict):
            logging.error("数据不是字典类型")
            return []
            
        if "Data" not in data:
            logging.error("数据中没有Data字段")
            return []
            
        data_list = data["Data"]
        if not isinstance(data_list, list):
            logging.error("Data字段不是列表类型")
            return []
            
        pids = [entry["pid"] for entry in data_list if isinstance(entry, dict) and "pid" in entry]
        if not pids:
            logging.warning("未找到任何有效的pid")
        return pids
    
    def save_data(self, data: Dict[str, Any], gid: int, day: str) -> str:
        """保存数据到文件"""
        filename = f"guild_data_{gid}_{day}.json"
        filepath = self.output_dir / filename
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return str(filepath)
    
    async def process_guild_data(self, gid: int, day: str) -> Dict[str, Any]:
        """异步处理公会数据"""
        logging.info(f"正在获取公会ID {gid} 在 {day} 的数据...")
        data = await self.fetch_guild_data(gid, day)
        
        if not data:
            return {"success": False, "message": "获取数据失败"}
        
        pids = self.extract_pids(data)
        if not pids:
            return {"success": False, "message": "未找到有效的pid数据"}
        
        # 获取公会名称
        gnick = "Unknown"
        if "Data" in data and len(data["Data"]) > 0:
            gnick = data["Data"][0].get("gnick", "Unknown")
        
        # 保存原始数据
        filepath = self.save_data(data, gid, day)
        
        # 获取PID详细信息
        pid_data = await self.fetch_pid_details(pids, self.pid_data_dir, gnick, gid)
        
        return {
            "success": True,
            "gid": gid,
            "day": day,
            "pids": pids,
            "pid_count": len(pids),
            "data_file": filepath,
            "gnick": gnick
        }

    async def fetch_pid_detail(self, pid: int) -> tuple[int, Dict[str, Any]]:
        """异步获取单个PID的详细信息"""
        url = f"{self.base_url}/pid_detail"
        params = {"pid": pid, "page": "1", "perPage": "50"}
        
        data = await self.fetch_with_retry(url, params)
        if not data:
            return pid, {"error": "获取数据失败"}
        return pid, data

    async def fetch_pid_details(self, pids: List[int], output_dir: Path, gnick: str = None, gid: int = None) -> Dict[str, Any]:
        """并发获取PID详细信息
        
        Args:
            pids (List[int]): PID列表
            output_dir (Path): 输出目录
            gnick (str): 公会名称
            gid (int): 公会ID
            
        Returns:
            Dict[str, Any]: 获取到的数据
        """
        if not pids:
            logging.warning("没有PID需要获取")
            return {}
            
        # 使用信号量限制并发请求数
        semaphore = asyncio.Semaphore(self.max_concurrent)
        
        async def fetch_with_semaphore(pid: int):
            async with semaphore:
                return await self.fetch_pid_detail(pid)
        
        # 并发获取所有PID的数据
        tasks = [fetch_with_semaphore(pid) for pid in pids]
        results = await asyncio.gather(*tasks)
        
        # 整理结果
        data_to_save = {pid: data for pid, data in results}
        
        # 保存数据
        if gnick and gid:
            # 使用公会名称和ID构建文件名，并保存到pid_data目录
            output_file = self.pid_data_dir / f"{gnick}_{gid}_pids_data.json"
        else:
            # 如果没有公会信息，使用通用文件名
            output_file = self.pid_data_dir / "hi20pids_data.json"
            
        with open(output_file, "w", encoding="utf-8") as json_file:
            json.dump(data_to_save, json_file, ensure_ascii=False, indent=4)
        
        logging.info(f"PID数据已保存到: {output_file}")
        return data_to_save

    async def fetch_all_guilds(self, day: str, wid: int = 1, ccid: int = 0, rank: str = "power", is_benfu: int = 1, is_quanfu: int = 0, page: int = 1, per_page: int = 500) -> Optional[Dict[str, Any]]:
        """获取全服联盟数据
        
        Args:
            day (str): 日期，格式为YYYYMMDD
            wid (int): 世界ID
            ccid (int): 国家ID
            rank (str): 排名类型
            is_benfu (int): 是否本服
            is_quanfu (int): 是否全服
            page (int): 页码
            per_page (int): 每页数量
            
        Returns:
            Optional[Dict[str, Any]]: 联盟数据，包含以下字段：
                - id: 记录ID
                - day: 日期
                - wid: 世界ID
                - ccid: 国家ID
                - gid: 公会ID
                - power: 公会战力
                - sname: 公会简称
                - fname: 公会全称
                - owner: 公会所有者
                - kil: 击杀数
                - di: 等级
                - c_power: 战力变化
                - c_kil: 击杀数变化
                - created_at: 创建时间
                - c_di: 等级变化
        """
        url = f"{self.base_url}/rank_guild"
        params = {
            "day": day,
            "wid": wid,
            "ccid": ccid,
            "rank": rank,
            "is_benfu": is_benfu,
            "is_quanfu": is_quanfu,
            "page": page,
            "perPage": per_page
        }
        
        data = await self.fetch_with_retry(url, params)
        if not data:
            logging.error(f"获取全服联盟数据失败")
            return None
            
        # 验证响应数据格式
        if not isinstance(data, dict) or "Code" not in data or "Data" not in data:
            logging.error("响应数据格式不正确")
            return None
            
        if data["Code"] != 0:
            logging.error(f"API返回错误: {data.get('Message', '未知错误')}")
            return None
            
        # 保存数据到文件
        filename = f"all_guilds_{day}.json"
        filepath = self.output_dir / filename
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        
        # 新增：保存为Excel（只保留指定字段）
        guilds_data = data["Data"]
        excel_path = self.output_dir / f"all_guilds_{day}.xlsx"
        # 只保留需要的字段，并设置中文表头和顺序
        export_fields = [
            ("wid", "区"),
            ("gid", "联盟id"),
            ("sname", "联盟简称"),
            ("fname", "联盟名称"),
            ("owner", "盟主"),
            ("power", "联盟总战力")
        ]
        filtered_guilds_data = [
            {field_cn: guild[field_en] for field_en, field_cn in export_fields if field_en in guild}
            for guild in guilds_data
        ]
        df = pd.DataFrame(filtered_guilds_data, columns=[field_cn for _, field_cn in export_fields])
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            ws = writer.sheets['Sheet1']

            # 设置表头样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill('solid', fgColor='4F81BD')
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置内容居中
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 自动调整列宽
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            # 加边框
            thin = Side(border_style='thin', color='CCCCCC')
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # 处理数据，添加一些统计信息
        total_power = sum(guild["power"] for guild in guilds_data)
        total_kills = sum(guild["kil"] for guild in guilds_data)
        avg_level = sum(guild["di"] for guild in guilds_data) / len(guilds_data) if guilds_data else 0
        
        return {
            "success": True,
            "data": data,
            "file_path": str(filepath),
            "excel_path": str(excel_path),
            "statistics": {
                "total_guilds": len(guilds_data),
                "total_power": total_power,
                "total_kills": total_kills,
                "average_level": round(avg_level, 2)
            }
        }

async def process_guild_data_async(fetcher: GuildDataFetcher, gid: int, dates: List[str]) -> List[Dict]:
    """异步处理多个日期的公会数据"""
    tasks = [fetcher.process_guild_data(gid, day) for day in dates]
    return await asyncio.gather(*tasks)

def get_date_range(start_date: str, days: int) -> List[str]:
    """生成日期范围"""
    start = datetime.strptime(start_date, "%Y%m%d")
    return [(start + timedelta(days=i)).strftime("%Y%m%d") for i in range(days)]

async def main_async():
    parser = argparse.ArgumentParser(description="战火公会数据获取工具")
    parser.add_argument("--gid", type=int, required=True, help="公会ID")
    parser.add_argument("--date", type=str, help="日期 (YYYYMMDD格式)，默认为今天")
    parser.add_argument("--days", type=int, default=1, help="获取天数，默认为1天")
    parser.add_argument("--output", type=str, help="输出目录，默认为'guild_data'")
    parser.add_argument("--max-concurrent", type=int, default=10, help="最大并发请求数")
    parser.add_argument("--max-retries", type=int, default=3, help="最大重试次数")
    parser.add_argument("--retry-delay", type=int, default=2, help="重试延迟（秒）")
    
    args = parser.parse_args()
    
    # 设置默认日期为今天
    if not args.date:
        args.date = datetime.now().strftime("%Y%m%d")
    
    # 创建数据获取器
    async with GuildDataFetcher(max_concurrent=args.max_concurrent) as fetcher:
        fetcher.max_retries = args.max_retries
        fetcher.retry_delay = args.retry_delay
        
        if args.output:
            fetcher.output_dir = Path(args.output)
            fetcher.output_dir.mkdir(exist_ok=True)
        
        # 获取日期范围
        dates = get_date_range(args.date, args.days)
        
        # 处理每个日期的数据
        results = await process_guild_data_async(fetcher, args.gid, dates)
        
        # 输出结果
        for result in results:
            if result["success"]:
                logging.info(f"\n处理完成:")
                logging.info(f"- 公会ID: {result['gid']}")
                logging.info(f"- 日期: {result['day']}")
                logging.info(f"- 找到PID数量: {result['pid_count']}")
                logging.info(f"- 数据文件: {result['data_file']}")
            else:
                logging.error(f"\n处理失败: {result['message']}")
        
        # 保存汇总报告
        if results:
            report_file = fetcher.output_dir / f"report_{args.gid}_{args.date}.json"
            with open(report_file, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=4)
            logging.info(f"\n汇总报告已保存到: {report_file}")

def main():
    # 配置日志
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('warpath_data.log'),
            logging.StreamHandler()
        ]
    )
    asyncio.run(main_async())

if __name__ == "__main__":
    main()
