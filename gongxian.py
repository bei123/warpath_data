import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import glob


input_files = glob.glob("D:/Desktop/冬日赛季/data/*.json") 
output_file = "D:/Desktop/冬日赛季/EX/gxtotal_rankings_with_server_alliance.xlsx"  


total_data_df = pd.DataFrame()


server_to_gnick = {
    "12": "阵营1", "21": "阵营1", "25": "阵营1",
    "18": "阵营2", "16": "阵营2", "3": "阵营2", "15": "阵营2"
}


for input_file in input_files:
    with open(input_file, "r", encoding="utf-8") as file:
        json_data = pd.read_json(file)

    if "Data" in json_data:
        data_df = pd.json_normalize(json_data["Data"])
    else:
        data_df = pd.DataFrame(json_data)

    
    data_df["kill_die_ratio"] = data_df["c_sumkill"] / data_df["c_die"]
    data_df["kill_die_ratio"] = data_df["kill_die_ratio"].replace([float("inf"), -float("inf")], 0).fillna(0)

    # 提取服务器和联盟名称（wid 和 gnick）
    data_df["wid"] = data_df["wid"].fillna("未知服务器") 
    data_df["gnick"] = data_df["gnick"].fillna("未知联盟")  

    # 将服务器的 wid 映射到对应的阵营
    data_df["gnick"] = data_df["wid"].map(server_to_gnick).fillna(data_df["gnick"])

    
    total_data_df = pd.concat([total_data_df, data_df], ignore_index=True)


total_data_df["contribution"] = total_data_df["c_sumkill"] / (total_data_df["c_die"] + 1e-5)  # 避免除以零


gnick_contrib = total_data_df.groupby("gnick").agg(
    total_kills=pd.NamedAgg(column="c_sumkill", aggfunc="sum"),
    total_deaths=pd.NamedAgg(column="c_die", aggfunc="sum"),
    total_contribution=pd.NamedAgg(column="contribution", aggfunc="sum")
)


total_data_df = total_data_df.merge(gnick_contrib["total_contribution"], on="gnick", how="left")
total_data_df["contribution_percentage"] = total_data_df["contribution"] / total_data_df["total_contribution"] * 100


total_data_df_sorted = total_data_df.sort_values(by=["gnick", "contribution"], ascending=[True, False])


total_data_df_sorted.to_excel(output_file, index=False)


workbook = load_workbook(output_file)
sheet = workbook.active


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        
        cell.border = thin_border
        
        cell.alignment = Alignment(horizontal="center", vertical="center")


workbook.save(output_file)

print(f"各阵营贡献度已成功导出到 {output_file}，并完成格式化。")
